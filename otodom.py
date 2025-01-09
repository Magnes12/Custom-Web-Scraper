import os
import sys
import requests
import openpyxl
import time
from otodom_locs import base_otodom_url_rent, base_otodom_url_sale, locations_otodom_rent, locations_otodom_sale
from bs4 import BeautifulSoup
from urllib.parse import urlencode


def create_excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.create_sheet(title="OTODOM WYNAJEM")
    sheet = wb["OTODOM WYNAJEM"]

    sheet['A1'] = 'Tytuł'
    sheet['B1'] = 'Cena [zł]'
    sheet['C1'] = 'Lokacja'
    sheet['D1'] = 'Powierzchnia [m2]'
    sheet['E1'] = 'URL'

    wb.save(file_name)
    return wb, sheet


def title(soup):
    title_element = soup.find_all('p', {'data-cy': 'listing-item-title', 'class': 'css-u3orbr e1g5xnx10'})
    print(f"Znaleziono {len(title_element)} ogłoszeń")
    return [title.text for title in title_element]


def price(soup):
    price_element = soup.find_all('span', {'class': 'css-2bt9f1 evk7nst0'})
    return [price.text.split('zł')[0] for price in price_element]


def loc(soup):
    localization_element = soup.find_all('p', {'class': 'css-42r2ms eejmx80'})
    return [loc.text for loc in localization_element]


def area(soup):
    # Find all the dl tags containing area information
    dl_elements = soup.find_all('dl', {'class': 'css-12dsp7a e1clni9t1'})

    areas = []
    for dl in dl_elements:

        dt = dl.find('dt', string="Powierzchnia")
        if dt:
            dd = dt.find_next('dd')
            if dd:
                area_text = dd.text.split()[0].replace(',', '.')
                areas.append(area_text)
    return areas


def url_site(soup):
    links_element = soup.find_all('a', {'data-cy': 'listing-item-link',
                                        'class': 'css-16vl3c1 e17g0c820'})
    return [f"https://www.otodom.pl{link.get('href')}" for link in links_element if link.get('href')]


def get_next_page(soup):
    pagination = soup.find('ul', {'data-cy': 'frontend.search.base-pagination.nexus-pagination'})
    if not pagination:
        print("Brak kolejnej strony. Zakończono przetwarzanie.\n")
        return False

    next_page = soup.find('li', {'aria-disabled': 'true',
                                 'title': 'Go to next Page'})
    if next_page:
        print("Brak kolejnej strony. Zakończono przetwarzanie.\n")
        return False

    return True


def scrape_otodom_data(wb, sheet, base_url, locations_otodom):
    
    row_number = 2

    print("OGŁOSZENIA OTODOM")

    for location, data in locations_otodom.items():
        page_number = 1
        while True:

            full_url = f"{base_url}{data['path']}?{urlencode(data['params'])}&page={page_number}"

            try:
                headers = {
                    'User-Agent': 'Mozilla/5.0',
                    'Accept-Language': 'en-US,en;q=0.9',
                    'Referer': 'https://www.google.com'
                }

                response = requests.get(full_url, headers=headers)
                response.raise_for_status()
            except requests.exceptions.RequestException as e:
                print(f"Błąd pobierania strony: {e}")
                sys.exit()

            soup = BeautifulSoup(response.text, 'html.parser')

            print(f"Lokalizacja: {location}")
            print(f"Strona: {page_number}")
            title_text = title(soup)
            price_text = price(soup)
            loc_text = loc(soup)
            area_text = area(soup)
            url_site_text = url_site(soup)

            if not (len(title_text) == len(price_text) == len(loc_text) == len(area_text) == len(url_site_text)):
                print("Error: Braki w danych")
                break

            for t, p, l, a, u in zip(title_text, price_text, loc_text, area_text, url_site_text):

                sheet[f'A{row_number}'].value = t

                sheet[f'B{row_number}'].value = p
                sheet[f'B{row_number}'].number_format = '#,##0.00 [$zł-415];[Red]-#,##0.00 [$zł-415]'

                sheet[f'C{row_number}'].value = l

                sheet[f'D{row_number}'].value = a
                sheet[f'D{row_number}'].number_format = '0'

                sheet[f'E{row_number}'].hyperlink = u
                sheet[f'E{row_number}'].style = "Hyperlink"

                row_number += 1

            if not get_next_page(soup):
                break

            page_number += 1
            time.sleep(1)


def otodom_main():
    excel_file = "LOKALE.xlsx"
    wb, sheet = create_excel(excel_file)

    scrape_otodom_data(wb, sheet, base_otodom_url_rent, locations_otodom_rent)
  
    sheet = wb.create_sheet(title="OTODOM SPRZEDAŻ")
    sheet = wb["OTODOM SPRZEDAŻ"]  
    
    scrape_otodom_data(wb, sheet, base_otodom_url_sale, locations_otodom_sale)


    wb.save(excel_file)
    print("Excel zapisany dla wszystkich lokalizacji OTODOM.\n")

    if sys.platform == "win32":
        print("Otwieram plik excel")
        os.startfile(excel_file)
